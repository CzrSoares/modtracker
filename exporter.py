"""
Export engine: generates PDF and Excel reports with embedded charts.
Uses matplotlib for chart rendering, reportlab for PDF, openpyxl for Excel.
"""
import io
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

# ── Palette ────────────────────────────────────────────────────────────────
C_BG      = "#07090e"
C_SURFACE = "#0d1018"
C_CARD    = "#111520"
C_TEXT    = "#dde3f0"
C_MUTED   = "#5a6482"
C_BAN     = "#ef4444"
C_MUTE    = "#f59e0b"
C_KICK    = "#8b5cf6"
C_WARN    = "#06b6d4"
C_MSG     = "#10b981"
C_PURPLE  = "#7c3aed"
C_CYAN    = "#22d3ee"

ACTION_COLORS = [C_BAN, C_MUTE, C_KICK, C_WARN]
ACTION_LABELS = ["Bans", "Mutes", "Kicks", "Warns"]


def _apply_dark_style(fig, ax):
    fig.patch.set_facecolor(C_BG)
    ax.set_facecolor(C_SURFACE)
    ax.tick_params(colors=C_MUTED, labelsize=8)
    for spine in ax.spines.values():
        spine.set_edgecolor("#1c2235")
    ax.xaxis.label.set_color(C_MUTED)
    ax.yaxis.label.set_color(C_MUTED)


def _fig_to_bytes(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    buf.seek(0)
    data = buf.read()
    buf.close()
    plt.close(fig)
    return data


# ── Chart builders ─────────────────────────────────────────────────────────

def build_bar_chart(agg: dict) -> bytes:
    rows = agg["rows"]
    if not rows:
        fig, ax = plt.subplots(figsize=(8, 3))
        _apply_dark_style(fig, ax)
        ax.text(0.5, 0.5, "No data", ha="center", va="center",
                color=C_MUTED, fontsize=12, transform=ax.transAxes)
        return _fig_to_bytes(fig)

    names = [r["name"] for r in rows]
    x = np.arange(len(names))
    width = 0.2
    fig, ax = plt.subplots(figsize=(max(8, len(names) * 1.4), 4.5))
    _apply_dark_style(fig, ax)

    fields = ["bans", "mutes", "kicks", "warns"]
    colors = ACTION_COLORS
    for i, (field, color, label) in enumerate(zip(fields, colors, ACTION_LABELS)):
        vals = [r[field] for r in rows]
        bars = ax.bar(x + i * width, vals, width, label=label,
                      color=color, alpha=0.85, zorder=3)
        for bar in bars:
            h = bar.get_height()
            if h > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, h + 0.08,
                        str(int(h)), ha="center", va="bottom",
                        color=C_TEXT, fontsize=7, fontweight="bold")

    ax.set_xticks(x + width * 1.5)
    ax.set_xticklabels(names, fontsize=9, color=C_TEXT, rotation=15 if len(names) > 5 else 0)
    ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
    ax.grid(axis="y", color="#1c2235", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    legend = ax.legend(framealpha=0.15, labelcolor=C_TEXT, fontsize=8,
                       facecolor=C_SURFACE, edgecolor="#1c2235")
    ax.set_title("Staff Actions Breakdown", color=C_TEXT, fontsize=11,
                 fontweight="bold", pad=12)
    fig.tight_layout()
    return _fig_to_bytes(fig)


def build_doughnut_chart(agg: dict) -> bytes:
    t = agg["totals"]
    values = [t["bans"], t["mutes"], t["kicks"], t["warns"], t["messages"]]
    labels = ["Bans", "Mutes", "Kicks", "Warns", "Messages"]
    colors = [C_BAN, C_MUTE, C_KICK, C_WARN, C_MSG]

    fig, ax = plt.subplots(figsize=(5, 4.5))
    _apply_dark_style(fig, ax)

    if sum(values) == 0:
        ax.text(0.5, 0.5, "No data", ha="center", va="center",
                color=C_MUTED, fontsize=12, transform=ax.transAxes)
        return _fig_to_bytes(fig)

    wedges, texts, autotexts = ax.pie(
        values, labels=None, colors=colors, autopct="%1.1f%%",
        startangle=90, pctdistance=0.78,
        wedgeprops={"linewidth": 2.5, "edgecolor": C_BG},
    )
    for at in autotexts:
        at.set_color(C_BG)
        at.set_fontsize(8)
        at.set_fontweight("bold")

    centre = plt.Circle((0, 0), 0.55, color=C_BG)
    ax.add_patch(centre)

    total_actions = t["total_actions"]
    ax.text(0, 0.08, str(total_actions), ha="center", va="center",
            color=C_TEXT, fontsize=18, fontweight="bold")
    ax.text(0, -0.18, "total actions", ha="center", va="center",
            color=C_MUTED, fontsize=7)

    patches = [mpatches.Patch(color=c, label=l) for c, l in zip(colors, labels)]
    ax.legend(handles=patches, loc="lower center", bbox_to_anchor=(0.5, -0.08),
              ncol=3, fontsize=8, framealpha=0.15, labelcolor=C_TEXT,
              facecolor=C_SURFACE, edgecolor="#1c2235")
    ax.set_title("Action Distribution", color=C_TEXT, fontsize=11,
                 fontweight="bold", pad=8)
    fig.tight_layout()
    return _fig_to_bytes(fig)


def build_messages_chart(agg: dict) -> bytes:
    rows = [r for r in agg["rows"] if r["messages"] > 0]
    if not rows:
        rows = agg["rows"]

    names = [r["name"] for r in rows]
    msgs  = [r["messages"] for r in rows]

    fig, ax = plt.subplots(figsize=(max(7, len(names) * 1.2), 3.8))
    _apply_dark_style(fig, ax)

    grad_colors = [C_MSG if m == max(msgs) else "#0d9e71" for m in msgs] if msgs else [C_MSG]
    bars = ax.bar(names, msgs, color=grad_colors, alpha=0.85, zorder=3, width=0.55)
    for bar in bars:
        h = bar.get_height()
        if h > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, h + max(msgs) * 0.01,
                    f"{int(h):,}", ha="center", va="bottom",
                    color=C_TEXT, fontsize=8, fontweight="bold")

    ax.set_xticks(range(len(names)))
    ax.set_xticklabels(names, fontsize=9, color=C_TEXT, rotation=15 if len(names) > 5 else 0)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{int(v):,}"))
    ax.grid(axis="y", color="#1c2235", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    ax.set_title("Messages per Staff Member", color=C_TEXT, fontsize=11,
                 fontweight="bold", pad=12)
    fig.tight_layout()
    return _fig_to_bytes(fig)


def build_stacked_chart(agg: dict) -> bytes:
    rows = agg["rows"]
    if not rows:
        fig, ax = plt.subplots(figsize=(8, 3))
        _apply_dark_style(fig, ax)
        ax.text(0.5, 0.5, "No data", ha="center", va="center",
                color=C_MUTED, fontsize=12, transform=ax.transAxes)
        return _fig_to_bytes(fig)

    names  = [r["name"] for r in rows]
    fields = ["bans", "mutes", "kicks", "warns"]
    colors = ACTION_COLORS

    fig, ax = plt.subplots(figsize=(max(8, len(names) * 1.3), 4))
    _apply_dark_style(fig, ax)

    bottoms = np.zeros(len(rows))
    for field, color, label in zip(fields, colors, ACTION_LABELS):
        vals = np.array([r[field] for r in rows], dtype=float)
        ax.bar(names, vals, bottom=bottoms, label=label,
               color=color, alpha=0.85, zorder=3, width=0.55)
        bottoms += vals

    ax.set_xticks(range(len(names)))
    ax.set_xticklabels(names, fontsize=9, color=C_TEXT, rotation=15 if len(names) > 5 else 0)
    ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
    ax.grid(axis="y", color="#1c2235", linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    ax.legend(framealpha=0.15, labelcolor=C_TEXT, fontsize=8,
              facecolor=C_SURFACE, edgecolor="#1c2235")
    ax.set_title("Stacked Actions per Staff", color=C_TEXT, fontsize=11,
                 fontweight="bold", pad=12)
    fig.tight_layout()
    return _fig_to_bytes(fig)


# ── PDF ────────────────────────────────────────────────────────────────────

def export_pdf(agg: dict, period_label: str) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Image, Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.8*cm, rightMargin=1.8*cm)

    # Colours
    bg      = HexColor(C_BG)
    surface = HexColor(C_SURFACE)
    text    = HexColor(C_TEXT)
    muted   = HexColor(C_MUTED)
    purple  = HexColor(C_PURPLE)
    cyan    = HexColor(C_CYAN)
    ban_c   = HexColor(C_BAN)
    mute_c  = HexColor(C_MUTE)
    kick_c  = HexColor(C_KICK)
    warn_c  = HexColor(C_WARN)
    msg_c   = HexColor(C_MSG)

    def style(name, **kw):
        defaults = dict(fontName="Helvetica", fontSize=10, textColor=text,
                        backColor=bg, leading=14)
        defaults.update(kw)
        return ParagraphStyle(name, **defaults)

    title_style   = style("title", fontSize=22, fontName="Helvetica-Bold",
                          textColor=purple, alignment=TA_CENTER, spaceAfter=4)
    sub_style     = style("sub",   fontSize=10, textColor=muted, alignment=TA_CENTER)
    section_style = style("sec",   fontSize=12, fontName="Helvetica-Bold",
                          textColor=cyan, spaceBefore=14, spaceAfter=6)
    body_style    = style("body",  fontSize=9,  textColor=text, leading=13)

    def chart_image(data: bytes, width_cm=15):
        img_buf = io.BytesIO(data)
        return Image(img_buf, width=width_cm*cm,
                     height=width_cm*cm*0.55)

    story = []

    # Header
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("ModTracker", title_style))
    story.append(Paragraph(f"Discord Staff Report — {period_label}", sub_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style))
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=HexColor("#1c2235")))
    story.append(Spacer(1, 0.4*cm))

    # KPI summary table
    t = agg["totals"]
    kpi_data = [
        ["BANS", "MUTES", "KICKS", "WARNS", "MESSAGES", "TOTAL ACTIONS"],
        [str(t["bans"]), str(t["mutes"]), str(t["kicks"]),
         str(t["warns"]), f"{t['messages']:,}", str(t["total_actions"])]
    ]
    kpi_table = Table(kpi_data, colWidths=[2.7*cm]*6)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), HexColor("#1c2235")),
        ("BACKGROUND",   (0,1), (-1,1), surface),
        ("TEXTCOLOR",    (0,0), (0,0),  ban_c),
        ("TEXTCOLOR",    (1,0), (1,0),  mute_c),
        ("TEXTCOLOR",    (2,0), (2,0),  kick_c),
        ("TEXTCOLOR",    (3,0), (3,0),  warn_c),
        ("TEXTCOLOR",    (4,0), (4,0),  msg_c),
        ("TEXTCOLOR",    (5,0), (5,0),  cyan),
        ("TEXTCOLOR",    (0,1), (-1,1), text),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME",     (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0), 7),
        ("FONTSIZE",     (0,1), (-1,1), 13),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[HexColor("#1c2235"), surface]),
        ("GRID",         (0,0), (-1,-1), 0.5, HexColor("#252d42")),
        ("ROUNDEDCORNERS", [4]),
        ("TOPPADDING",   (0,0), (-1,-1), 7),
        ("BOTTOMPADDING",(0,0), (-1,-1), 7),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5*cm))

    # Charts
    story.append(Paragraph("Actions Breakdown", section_style))
    story.append(chart_image(build_bar_chart(agg), width_cm=16))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Stacked Actions", section_style))
    story.append(chart_image(build_stacked_chart(agg), width_cm=16))
    story.append(Spacer(1, 0.4*cm))

    # Doughnut + messages side by side
    story.append(Paragraph("Distribution & Messages", section_style))
    donut_bytes = build_doughnut_chart(agg)
    msg_bytes   = build_messages_chart(agg)
    side_data = [[chart_image(donut_bytes, width_cm=7.8),
                  chart_image(msg_bytes,   width_cm=7.8)]]
    side_table = Table(side_data, colWidths=[8.3*cm, 8.3*cm])
    side_table.setStyle(TableStyle([
        ("ALIGN",  (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(side_table)
    story.append(Spacer(1, 0.5*cm))

    # Rankings table
    story.append(Paragraph("Staff Rankings", section_style))
    header = ["#", "Staff Member", "Bans", "Mutes", "Kicks", "Warns",
              "Messages", "Actions", "% Share"]
    rows_data = [header]
    total_act = t["total_actions"] or 1
    for i, r in enumerate(agg["rows"], 1):
        pct = f"{r['total_actions']/total_act*100:.1f}%"
        rows_data.append([str(i), r["name"], str(r["bans"]), str(r["mutes"]),
                          str(r["kicks"]), str(r["warns"]),
                          f"{r['messages']:,}", str(r["total_actions"]), pct])
    # Totals row
    rows_data.append(["—", "TOTAL", str(t["bans"]), str(t["mutes"]),
                      str(t["kicks"]), str(t["warns"]),
                      f"{t['messages']:,}", str(t["total_actions"]), "100%"])

    col_widths = [0.8*cm, 4.2*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2.2*cm, 1.8*cm, 1.8*cm]
    rank_table = Table(rows_data, colWidths=col_widths, repeatRows=1)
    rank_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  HexColor("#1c2235")),
        ("TEXTCOLOR",     (0,0), (-1,0),  cyan),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("ALIGN",         (1,0), (1,-1),  "LEFT"),
        ("ROWBACKGROUNDS",(0,1), (-1,-2), [surface, HexColor("#0f1420")]),
        ("BACKGROUND",    (0,-1),(-1,-1), HexColor("#1c2235")),
        ("TEXTCOLOR",     (0,-1),(-1,-1), cyan),
        ("FONTNAME",      (0,-1),(-1,-1), "Helvetica-Bold"),
        ("GRID",          (0,0), (-1,-1), 0.4, HexColor("#252d42")),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("TEXTCOLOR",     (0,1), (-1,-2), text),
    ]))
    story.append(rank_table)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── Excel ──────────────────────────────────────────────────────────────────

def export_excel(agg: dict, period_label: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = period_label[:31]
    ws.sheet_view.showGridLines = False

    # Helpers
    def hex_fill(hex_str):
        return PatternFill("solid", fgColor=hex_str.lstrip("#"))

    def thin_border(color="252d42"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def write(row, col, value, bold=False, size=10, color="dde3f0",
              bg=None, align="left", italic=False, border=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name="Calibri", bold=bold, size=size,
                         color=color.lstrip("#"), italic=italic)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg:
            cell.fill = hex_fill(bg)
        if border:
            cell.border = thin_border()
        return cell

    # Column widths
    col_widths = [5, 28, 10, 10, 10, 10, 16, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    for r in range(1, 6):
        ws.row_dimensions[r].height = 22

    # Title block
    ws.merge_cells("A1:I1")
    write(1, 1, "ModTracker — Discord Staff Report", bold=True, size=16,
          color=C_PURPLE.lstrip("#"), bg=C_BG.lstrip("#"), align="center")
    ws.merge_cells("A2:I2")
    write(2, 1, f"Period: {period_label}", size=11, color=C_CYAN.lstrip("#"),
          bg=C_BG.lstrip("#"), align="center")
    ws.merge_cells("A3:I3")
    write(3, 1, f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
          size=9, color="5a6482", bg=C_BG.lstrip("#"), align="center", italic=True)
    ws.row_dimensions[4].height = 8

    # KPI row
    t = agg["totals"]
    kpi_items = [
        ("BANS",    t["bans"],            C_BAN),
        ("MUTES",   t["mutes"],           C_MUTE),
        ("KICKS",   t["kicks"],           C_KICK),
        ("WARNS",   t["warns"],           C_WARN),
        ("MESSAGES",t["messages"],        C_MSG),
        ("TOTAL ACTIONS", t["total_actions"], C_CYAN),
    ]
    # Merge pairs for KPI display
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 30

    for idx, (label, val, color) in enumerate(kpi_items):
        col = idx + 1
        write(5, col, label, bold=True, size=8, color=color.lstrip("#"),
              bg="1c2235", align="center", border=True)
        write(6, col, val, bold=True, size=14, color=color.lstrip("#"),
              bg="111520", align="center", border=True)

    ws.row_dimensions[7].height = 8

    # Table header row 8
    headers = ["#", "Staff Member", "Bans", "Mutes", "Kicks", "Warns",
               "Messages", "Total Actions", "% of Total"]
    ws.row_dimensions[8].height = 20
    for col, h in enumerate(headers, 1):
        write(8, col, h, bold=True, size=9, color=C_CYAN.lstrip("#"),
              bg="1c2235", align="center", border=True)

    # Data rows
    total_act = t["total_actions"] or 1
    for i, r in enumerate(agg["rows"], 1):
        row = 8 + i
        ws.row_dimensions[row].height = 18
        pct = f"{r['total_actions']/total_act*100:.1f}%"
        row_bg = "111520" if i % 2 == 1 else "0f1420"
        vals = [i, r["name"], r["bans"], r["mutes"], r["kicks"],
                r["warns"], r["messages"], r["total_actions"], pct]
        colors_row = ["8892aa", "dde3f0", "fca5a5", "fcd34d", "c4b5fd",
                      "67e8f9", "6ee7b7", "dde3f0", "dde3f0"]
        for col, (val, color) in enumerate(zip(vals, colors_row), 1):
            write(row, col, val, size=9, color=color, bg=row_bg,
                  align="center" if col != 2 else "left", border=True)

    # Totals row
    tot_row = 8 + len(agg["rows"]) + 1
    ws.row_dimensions[tot_row].height = 20
    tot_vals = ["—", "TOTAL", t["bans"], t["mutes"], t["kicks"],
                t["warns"], t["messages"], t["total_actions"], "100%"]
    for col, val in enumerate(tot_vals, 1):
        write(tot_row, col, val, bold=True, size=9, color=C_CYAN.lstrip("#"),
              bg="1c2235", align="center" if col != 2 else "left", border=True)

    # Charts sheet
    ws_charts = wb.create_sheet("Charts")
    ws_charts.sheet_view.showGridLines = False
    ws_charts.sheet_properties.tabColor = "7c3aed"

    # Background fill
    for row in range(1, 60):
        ws_charts.row_dimensions[row].height = 15

    chart_items = [
        ("Actions Breakdown", build_bar_chart(agg)),
        ("Stacked Actions",   build_stacked_chart(agg)),
        ("Distribution",      build_doughnut_chart(agg)),
        ("Messages",          build_messages_chart(agg)),
    ]

    positions = ["A1", "A26", "M1", "M26"]
    ws_charts.column_dimensions["A"].width = 2
    for i in range(1, 30):
        ws_charts.column_dimensions[get_column_letter(i)].width = 4

    for (title, chart_bytes), pos in zip(chart_items, positions):
        img_buf = io.BytesIO(chart_bytes)
        xl_img = XLImage(img_buf)
        xl_img.width  = 480
        xl_img.height = 270
        ws_charts.add_image(xl_img, pos)

    # Set background colour workaround (column fills)
    ws.sheet_properties.tabColor = "7c3aed"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
